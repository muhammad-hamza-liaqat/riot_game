import openpyxl
from django.http import HttpResponse

def generate_match_excel(response_data, summoner_only=False, selected_match_id=None):
    matches = response_data['matches']
    if selected_match_id:
        matches = [m for m in matches if m['match_id'] == selected_match_id]
        if not matches:
            return None, {'error': 'Selected match not found.'}

    wb = openpyxl.Workbook()

    wb.remove(wb.active)

    ws_summoner_stats = wb.create_sheet(title="Summoner Stats")
    ws_summoner_stats.append([
        'Summoner Name', 'Total Matches', 'Average KDA', 'Average Gold',
        'Average Damage', 'Average CS', 'Win Rate'
    ])
    summoner_stats = response_data['summoner_stats']
    ws_summoner_stats.append([
        response_data['summoner_name'],
        summoner_stats['total_matches'],
        summoner_stats['avg_kda'],
        summoner_stats['avg_gold'],
        summoner_stats['avg_damage'],
        summoner_stats['avg_cs'],
        f"{summoner_stats['win_rate']:.1f}%"
    ])

    ws_per_match_stats = wb.create_sheet(title="Per-Match Stats")
    ws_per_match_stats.append([
        'Match ID', 'Date', 'Champion', 'KDA', 'CS', 'Gold', 'Damage', 'Win'
    ])
    for match in response_data['matches']:  # Use all matches for Per-Match Stats
        for participant in match['match_data']['info']['participants']:
            if participant.get('puuid') == response_data['puuid']:
                ws_per_match_stats.append([
                    match['match_id'],
                    match['match_date'],
                    participant['championName'],
                    f"{participant['kills']}/{participant['deaths']}/{participant['assists']}",
                    participant['totalMinionsKilled'] + participant['neutralMinionsKilled'],
                    participant['goldEarned'],
                    participant['totalDamageDealtToChampions'],
                    'Yes' if participant.get('win', False) else 'No'
                ])

    ws_match_details = wb.create_sheet(title="Match Details")
    ws_match_details.append([
        'Match ID', 'Date', 'Duration (min)', 'Game Mode', 'Queue', 'Version', 'Map'
    ])
    for match in matches:
        queue_id = match['match_data']['info']['queueId']
        queue_type = (
            'Normal Draft' if queue_id == 400 else
            'Ranked Solo' if queue_id == 420 else
            'Ranked Flex' if queue_id == 440 else str(queue_id)
        )
        ws_match_details.append([
            match['match_id'],
            match['match_date'],
            f"{match['game_duration_minutes']:.1f}",
            match['match_data']['info']['gameMode'],
            queue_type,
            match['match_data']['info']['gameVersion'],
            "Summoner's Rift"
        ])

    ws_team_stats = wb.create_sheet(title="Team Stats")
    ws_team_stats.append([
        'Match ID', 'Team ID', 'Result', 'Kills', 'Turrets', 'Gold', 'Dragons', 'Barons', 'Heralds'
    ])
    for match in matches:
        for team in match['match_data']['info']['teams']:
            ws_team_stats.append([
                match['match_id'],
                team['teamId'],
                'Victory' if team['win'] else 'Defeat',
                team['objectives']['champion']['kills'],
                team['objectives']['tower']['kills'],
                match['team_totals'].get(team['teamId'], {}).get('total_gold', 'N/A'),
                team['objectives']['dragon']['kills'],
                team['objectives']['baron']['kills'],
                team['objectives']['riftHerald']['kills']
            ])

    ws_participant_stats = wb.create_sheet(title="Participant Stats")
    headers = [
        'Match ID', 'Summoner', 'Champion', 'Position', 'Level', 'KDA', 'CS', 'Gold',
        'Damage to Champions', 'Physical Damage', 'Magic Damage', 'True Damage',
        'Damage Taken', 'Healing Done', 'Shielding Done', 'Vision Score',
        'Wards Placed', 'Wards Killed', 'CC Score', 'Dragon Kills', 'Baron Kills',
        'Turret Kills', 'Win', 'Summoner Spells', 'Runes', 'Team ID', 'Items'
    ]
    ws_participant_stats.append(headers)
    for match in matches:
        for participant in match['match_data']['info']['participants']:
            if summoner_only and participant.get('puuid') != response_data['puuid']:
                continue

            items = [
                f"Item{i}: {participant.get(f'item{i}', 0)}"
                for i in range(7) if participant.get(f'item{i}', 0) != 0
            ]
            items_str = '; '.join(items) if items else 'None'

            summoner_spells = f"{participant.get('summoner1Id', 'N/A')} & {participant.get('summoner2Id', 'N/A')}"

            perks = participant.get('perks', {}).get('styles', [])
            runes = []
            for style in perks:
                for selection in style.get('selections', []):
                    runes.append(str(selection.get('perk', '')))
            runes_str = ', '.join(runes) if runes else 'None'

            row = [
                match['match_id'],
                response_data['summoner_name'] + ' (YOU)' if participant.get('puuid') == response_data.get('puuid') else participant.get('summonerName', 'Unknown'),
                participant['championName'],
                participant['teamPosition'],
                participant['champLevel'],
                f"{participant['kills']}/{participant['deaths']}/{participant['assists']}",
                participant['totalMinionsKilled'] + participant['neutralMinionsKilled'],
                participant['goldEarned'],
                participant['totalDamageDealtToChampions'],
                participant['physicalDamageDealtToChampions'],
                participant['magicDamageDealtToChampions'],
                participant['trueDamageDealtToChampions'],
                participant['totalDamageTaken'],
                participant.get('totalHeal', 0),
                participant.get('totalShieldingOnTeammates', 0),
                participant['visionScore'],
                participant.get('wardsPlaced', 0),
                participant.get('wardsKilled', 0),
                participant.get('timeCCingOthers', 0),
                participant.get('dragonKills', 0),
                participant.get('baronKills', 0),
                participant.get('turretKills', 0),
                'Yes' if participant.get('win', False) else 'No',
                summoner_spells,
                runes_str,
                participant['teamId'],
                items_str
            ]
            ws_participant_stats.append(row)

    ws_team_100_stats = wb.create_sheet(title="Team 100 Detailed Stats")
    ws_team_100_stats.append(['Match ID', 'Summoner', 'Damage', 'Team Damage %', 'Vision Score', 'Vision/Min', 'Objectives (D/H/B)'])
    for match in matches:
        for participant in match['match_data']['info']['participants']:
            if participant.get('teamId') == 100:
                ws_team_100_stats.append([
                    match['match_id'],
                    response_data['summoner_name'] if participant.get('puuid') == response_data['puuid'] else participant.get('summonerName', 'Unknown'),
                    participant['totalDamageDealtToChampions'],
                    f"{participant['challenges']['teamDamagePercentage']:.1f}%",
                    participant['visionScore'],
                    f"{participant['challenges']['visionScorePerMinute']:.1f}",
                    f"D:{participant.get('dragonKills', 0)} / H:{participant['challenges'].get('voidMonsterKill', 0)} / B:{participant.get('baronKills', 0)}"
                ])

    ws_team_200_stats = wb.create_sheet(title="Team 200 Detailed Stats")
    ws_team_200_stats.append(['Match ID', 'Summoner', 'Damage', 'Team Damage %', 'Vision Score', 'Vision/Min', 'Objectives (D/H/B)'])
    for match in matches:
        for participant in match['match_data']['info']['participants']:
            if participant.get('teamId') == 200:
                ws_team_200_stats.append([
                    match['match_id'],
                    response_data['summoner_name'] if participant.get('puuid') == response_data['puuid'] else participant.get('summonerName', 'Unknown'),
                    participant['totalDamageDealtToChampions'],
                    f"{participant['challenges']['teamDamagePercentage']:.1f}%",
                    participant['visionScore'],
                    f"{participant['challenges']['visionScorePerMinute']:.1f}",
                    f"D:{participant.get('dragonKills', 0)} / H:{participant['challenges'].get('voidMonsterKill', 0)} / B:{participant.get('baronKills', 0)}"
                ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="matches_{response_data["summoner_name"]}.xlsx"'
    wb.save(response)
    return response, None